import re
import json
from ipdb import set_trace
import os
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
from openpyxl.utils import get_column_letter

import pandas as pd
from openpyxl import Workbook

def extract_vulnerabilities(content, report_path, vul_record):
    pattern = re.compile(
        r'Vulnerability (\d+):.*?'
        r'Source.*?File: (.*?), Function: (.*?)\s+pc: (-?\d+), trigger: "(.*?)".*?'
        r'Sink.*?File: (.*?), Function: (.*?)\s+pc: (-?\d+), trigger: "(.*?)"',
        re.DOTALL
    )
    # set_trace()
    for match in re.finditer(pattern, content):
        vuln_id = match.group(1)
        source = {
            "file": os.path.basename(match.group(2).strip()),
            "function": match.group(3).strip(),
            "pc": int(match.group(4)),
            "trigger": match.group(5)
        }
        sink = {
            "file": os.path.basename(match.group(6).strip()),
            "function": match.group(7).strip(),
            "pc": int(match.group(8)),
            "trigger": match.group(9),
        }
        result = {
            "vulnerability_id": vuln_id,
            "source": source,
            "sink": sink,
            "report_path": report_path
        }
        source_tuple = tuple(sorted(source.items()))
        sink_tuple = tuple(sorted(sink.items()))
        key = (source_tuple, sink_tuple)
        vul_record[key].append(result)


def travel_vul_report(vul_path):
    vul_record = defaultdict(list)
    # 使用 os.walk 遍历目录树
    for root, dirs, files in os.walk(vul_path):
        for file in files:
            if ":True:" in file:
                report_file = os.path.join(root, file)
                report_content = None
                # set_trace()
                max_size = 10 * 1024 * 1024
                with open(report_file) as f:
                    report_content = f.read(max_size)
                print(report_file)
                if report_content:
                    extract_vulnerabilities(report_content, report_file, vul_record)
    return vul_record


def flatten_vulnerability(data, vul_type):
    """展平嵌套的漏洞数据结构"""
    flattened = {
        'vuln_type': vul_type,
        'vuln_id': data['vulnerability_id'],
        'source_file': data['source']['file'],
        'source_function': data['source']['function'],
        'source_pc': data['source']['pc'],
        'source_trigger': data['source']['trigger'],
        'sink_file': data['sink']['file'],
        'sink_function': data['sink']['function'],
        'sink_pc': data['sink']['pc'],
        'sink_trigger': data['sink']['trigger'],
        'report_path': data['report_path']
    }
    return flattened

def format_excel(writer):
    """使用openpyxl正确的样式设置方法"""
    workbook = writer.book
    sheet = writer.sheets['Sheet1']

    # 创建并注册标题样式
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(bold=True, color="FFFFFF")
    header_style.fill = PatternFill("solid", fgColor="4F81BD")
    header_style.alignment = Alignment(horizontal="center")
    workbook.add_named_style(header_style)

    # 应用标题样式
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=1, column=col)
        cell.style = "header_style"

    # 设置列宽（示例）
    column_widths = {
        'A': 8,    # vul_type
        'B': 15,   # vuln_id
        'C': 60,   # source_file
        'D': 25,   # source_function
        'E': 10,   # source_pc
        'F': 25,   # source_trigger
        'G': 60,   # sink_file
        'H': 25,   # sink_function
        'I': 10,   # sink_pc
        'J': 25,   # sink_trigger
        'K': 50    # report_path
    }

    for col_letter, width in column_widths.items():
        sheet.column_dimensions[col_letter].width = width

def classify_vul_into_xlsx(vul_record, xlsx_path):
    # 处理数据并创建DataFrame
    processed_data = []
    vul_type = 0
    for result_json_list in vul_record.values():
        vul_type = vul_type + 1
        for result_json in result_json_list:
            processed_data.append(flatten_vulnerability(result_json, vul_type))

    df = pd.DataFrame(processed_data)

    # 保存文件时调用
    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        format_excel(writer)  # 应用格式
    print(f"Excel文件已生成: {xlsx_path}")


def vul_brand_path(report_dir, xlsx_path):
    vul_record = travel_vul_report(report_dir)
    classify_vul_into_xlsx(vul_record, xlsx_path)
